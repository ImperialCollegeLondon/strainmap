from pytest import approx, mark, raises
from unittest.mock import MagicMock, patch


def test_add_metadata(strainmap_data):
    from strainmap.models.writers import add_metadata
    import openpyxl as xlsx

    dataset = strainmap_data.data_files.datasets[0]

    wb = xlsx.Workbook()
    ws = wb.create_sheet("Parameters")
    assert "Parameters" in wb.sheetnames

    add_metadata(strainmap_data.metadata(dataset), ws)
    assert ws.max_row == 8


@mark.skip(reason="Refactoring of export to excel in progress")
def test_add_markers(markers):
    from strainmap.models.writers import add_markers
    import numpy as np
    import openpyxl as xlsx

    wb = xlsx.Workbook()
    ws = wb.create_sheet("Parameters")

    colnames = (
        "Parameter",
        "Region",
        "P",
        "S",
        "PSS",
        "P",
        "S",
        "PSS",
        "ES",
        "P",
        "S",
        "PSS",
    )

    p = ("Frame", "Strain (%)", "Time (s)")
    region_names = "AS", "A", "AL", "IL", "I", "IS"
    add_markers(
        markers[None, :, :, :],
        ws,
        colnames=colnames,
        p=p,
        region_names=region_names,
        title="Global",
    )
    m = markers[None, :, :, :].transpose((3, 0, 1, 2)).reshape((-1, 12))
    expected = 5 + len(m)
    assert ws.max_row == expected
    assert ws["A5"].value == "Parameter"
    assert ws["L8"].value == approx(np.around(markers[2, 2, 2], 3))


@mark.skip(reason="Refactoring of export to excel in progress")
def test_add_velocity(velocity):
    from strainmap.models.writers import add_velocity
    import numpy as np
    import openpyxl as xlsx

    wb = xlsx.Workbook()
    ws = wb.create_sheet("Velocity")
    region_names = "AS", "A", "AL", "IL", "I", "IS"

    add_velocity(velocity[None, :, :], region_names, ws)
    assert ws.max_row == 52
    assert ws["A52"].value == approx(np.around(velocity[0, -1], 3))
    assert ws["C52"].value == approx(np.around(velocity[1, -1], 3))
    assert ws["E52"].value == approx(np.around(velocity[2, -1], 3))


def test_save_group(tmp_path):
    from strainmap.models.writers import save_group

    to_netcdf = MagicMock()

    with patch("xarray.DataArray.to_netcdf", to_netcdf):
        import xarray as xr

        da = xr.DataArray([1, 2, 3], dims=["x"], coords={"x": ["a", "b", "c"]})

        filename = tmp_path / "my_data.nc"
        save_group(filename, da, "speed")
        to_netcdf.assert_called_with(filename, mode="w", group="speed", encoding={})

        to_netcdf.reset_mock()
        filename.open("w").close()
        save_group(filename, da, "speed")
        to_netcdf.assert_called_with(filename, mode="a", group="speed", encoding={})

        to_netcdf.reset_mock()
        save_group(filename, da, "speed", overwrite=True)
        to_netcdf.assert_called_with(filename, mode="w", group="speed", encoding={})

        to_netcdf.reset_mock()
        save_group(filename, da, "speed", to_int=True)
        assert to_netcdf.call_args[-1]["encoding"] != {}


def test_save_attribute(tmp_path):
    from strainmap.models.writers import save_attribute
    import h5py

    filename = tmp_path / "my_data.nc"
    save_attribute(filename, name="Thor")

    f = h5py.File(filename, mode="r")
    assert f.attrs["name"] == "Thor"


def test_write_netcdf_file(tmp_path):
    from strainmap.models.writers import write_netcdf_file
    import xarray as xr
    import numpy as np
    import h5py

    ds = xr.Dataset(
        {"foo": (("x", "y"), np.random.rand(4, 5)), "bar": (("x"), np.random.rand(4))},
        coords={
            "x": [10, 20, 30, 40],
            "y": np.linspace(0, 10, 5),
            "z": ("y", list("abcde")),
        },
    )

    filename = tmp_path / "my_data.nc"
    write_netcdf_file(filename, foo=ds.foo, bar=ds.bar, patient="Thor", age=38)

    data = h5py.File(filename, mode="r")
    assert data.attrs["patient"] == "Thor"
    assert data.attrs["age"] == 38

    for var in (ds.foo, ds.bar):
        assert var.name in data
        assert var.name in data[var.name]
        for coord in var.coords:
            assert coord in data[var.name]


def test_repack_file(tmp_path):
    from strainmap.models.writers import write_netcdf_file, repack_file
    import xarray as xr

    da = xr.DataArray([1, 2, 3], dims=["x"], coords={"x": ["a", "b", "c"]})

    filename = tmp_path / "my_data.nc"

    # Write the file twice to create some space issues
    write_netcdf_file(filename, foo=da)
    write_netcdf_file(filename, foo=da)

    repack_file(filename)
    target = filename.parent / f"~{filename.name}"
    assert target.is_file()
    assert target.stat().st_size < filename.stat().st_size

    with raises(RuntimeError):
        repack_file(filename, target)

    target = filename.parent / "new_filename.nc"
    repack_file(filename, target)
    assert target.is_file()
